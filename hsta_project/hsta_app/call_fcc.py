import openpyxl as xl
import datetime
import zipfile
from .hsta import Fcc
from django.conf import settings
from .models import HstaData
from shutil import copy2
import shutil
from zipfile import ZipFile
import os


def calling_fcc_func(pk, zip_file_name):
    data = HstaData.objects.filter(id=pk)
    for Data in data:
        isbn = Data.ISBN
        author = Data.author
        title = Data.title
        chapter = Data.chapter
        edition = Data.edition
        zip_file_path = Data.zip_file.path

    dir_path = os.path.join(settings.MEDIA_ROOT, 'ID' + str(pk))
    os.mkdir(dir_path)
    with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
        zip_ref.extractall(dir_path)

    file_name = str(zip_file_name).rsplit('.', 1)[0]

    print("------------------------------")
    print(dir_path)
    print(file_name)
    print("------------------------------")
    c2 = Fcc()

    if len(os.listdir(r"" + dir_path + '\\' + file_name + '\\Manuscripts\\')) == 0:
        File = "None"
    else:
        for file in os.listdir(r"" + dir_path + '\\' + file_name + '\\Manuscripts\\'):
            if file.endswith(".docx"):
                File = file

    print("---->", File)
    figurePath = r"" + dir_path + '\\' + file_name + '\\Figures\\'
    oldallImagesList = os.listdir(r"{}".format(figurePath))

    try:
        c2.Image_auto(figurePath, isbn)
    except:
        copy2(r"" + settings.MEDIA_ROOT + '\\excel_file_template\\HS_TA_Template.xlsx',
              r"" + settings.MEDIA_ROOT + '\\created_report\\')
        x = 0
        filePathxl = r"" + settings.MEDIA_ROOT + "\created_report\HS_TA_Template.xlsx"
        wb = xl.load_workbook(filePathxl)
        sheet = wb.active
        for i in range(1, 56):
            if sheet.cell(i, 1).value == "TOTALS":
                x = i
                break
        today = datetime.date.today()
        date = "{:%d-%b-%Y}".format(today)
        sheet.cell(2, 2).value = author
        sheet.cell(3, 2).value = title
        sheet.cell(4, 9).value = chapter
        sheet.cell(2, 9).value = isbn
        sheet.cell(5, 9).value = date
        sheet.cell(5, 2).value = "xyz@aptaracorp.com"
        sheet.cell(3, 9).value = edition

    copy2(r"" + settings.MEDIA_ROOT + '\\excel_file_template\\HS_TA_Template.xlsx',
          r"" + settings.MEDIA_ROOT + '\\created_report\\')

    docPath = r"" + dir_path + '\\' + file_name + '\\Manuscripts\\' + File
    print("------DOC PATH---->", docPath)

    if File != "None":
        try:
            c2.createReport(figurePath, r"" + docPath, oldallImagesList, title, chapter, isbn, edition, author)
        except:
            x = 0
            filePathxl = r"" + settings.MEDIA_ROOT + "\created_report\HS_TA_Template.xlsx"
            wb = xl.load_workbook(filePathxl)
            sheet = wb.active
            for i in range(1, 56):
                if sheet.cell(i, 1).value == "TOTALS":
                    x = i
                    break
            today = datetime.date.today()
            date = "{:%d-%b-%Y}".format(today)
            sheet.cell(2, 2).value = author
            sheet.cell(3, 2).value = title
            sheet.cell(4, 9).value = chapter
            sheet.cell(2, 9).value = isbn
            sheet.cell(5, 9).value = date
            sheet.cell(5, 2).value = "xyz@aptaracorp.com"
            sheet.cell(3, 9).value = edition

            wb.save(filePathxl)
    else:
        x = 0
        filePathxl = r"" + settings.MEDIA_ROOT + "\created_report\HS_TA_Template.xlsx"
        wb = xl.load_workbook(filePathxl)
        sheet = wb.active
        for i in range(1, 56):
            if sheet.cell(i, 1).value == "TOTALS":
                x = i
                break
        today = datetime.date.today()
        date = "{:%d-%b-%Y}".format(today)
        sheet.cell(2, 2).value = author
        sheet.cell(3, 2).value = title
        sheet.cell(4, 9).value = chapter
        sheet.cell(2, 9).value = isbn
        sheet.cell(5, 9).value = date
        sheet.cell(5, 2).value = "xyz@aptaracorp.com"
        sheet.cell(3, 9).value = edition

        wb.save(filePathxl)

    copy2(r"" + settings.MEDIA_ROOT + '\\created_report\\HS_TA_Template.xlsx',
          dir_path + '\\' + file_name)

    shutil.rmtree(dir_path + '\\' + file_name + '\\Manuscripts')
    os.remove(r"" + settings.MEDIA_ROOT + '\\created_report\\HS_TA_Template.xlsx')

    outputPath = dir_path + '\\' + file_name
    return outputPath
